
import openpyxl # excel操作のライブラリ

wb_summary = openpyxl.load_workbook("C:\出社在宅集計表_まとめ.xlsx")

department_list = ["人事部", "企画部", "営業部"]
month_list = ["4月", "5月", "6月"]

for department in department_list:
    wb_department = openpyxl.load_workbook("C:\出社在宅集計表_{}.xlsx".format(department))

    if department == "人事部":
            syussya_row = 2
            zaitaku_row = 3
    if department == "企画部":
            syussya_row = 4
            zaitaku_row = 5
    if department == "営業部":
            syussya_row = 6
            zaitaku_row = 7

    for month in month_list:
        ws_department = wb_department[month]
        ws_summary = wb_summary[month]

        for i in range(ws_department.max_column-1):
            ws_summary.cell(row=syussya_row, column=3+i).value = ws_department.cell(row=2, column=2+i).value
            ws_summary.cell(row=zaitaku_row, column=3+i).value = ws_department.cell(row=3, column=2+i).value
    
wb_summary.save("出社在宅集計表_まとめ2.xlsx")



